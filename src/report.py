from openpyxl import Workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment, Font
from datetime import date, datetime, timedelta

import os

from excel_pdf import generate_pdf

class cash_deposit_struct:
    def __init__(self, date:tuple, amount:tuple, deposit_data:dict, cash_list: list):
        self.banking_date = date[0]
        self.start_date = date[1]
        self.end_date = date[2]

        self.cash_amount = amount[0]
        self.coin_amount = amount[1]
        self.total_amount = amount[2]

        self.deposit_data = deposit_data

        self.total_cash_amount = {}

        self.cash_list = cash_list


def generate_report(data:cash_deposit_struct):
    wb = Workbook()
    #Create a new worksheet
    banking_date= datetime.strptime(data.banking_date,"%A, %d/%m/%Y").strftime('%d%b%y')
    summary_title = f"Summary {banking_date}"

    #Create worksheet for the daily cash amount
    output_ws = wb.active
    output_ws.title = summary_title
    output_ws['A1'] = "Cash Deposit Report Summary"
    output_ws.column_dimensions['A'].width = 17
    output_ws.column_dimensions['B'].width = 11
    output_ws['A1'].font = Font(bold=True)

    #Add banking details
    output_ws['A3'] = "Banking date"
    output_ws['B3'] = data.banking_date
    output_ws['B3'].number_format = 'dd/mm/yyyy'
    output_ws['A4'] = "Start period"
    output_ws['B4'] = data.start_date
    output_ws['B4'].number_format = 'dd/mm/yyyy'
    output_ws['A5'] = "End period"
    output_ws['B5'] = data.end_date
    output_ws['B5'].number_format = 'dd/mm/yyyy'

    output_ws['E3'] = "Cash amount"
    output_ws['G3'] = data.cash_amount
    output_ws['G3'].number_format =  u'"$ "#,##0.00'
    output_ws['E4'] = "Coin amount"
    output_ws['G4'] = data.coin_amount
    output_ws['G4'].number_format =  u'"$ "#,##0.00'
    output_ws['E5'] = "Total amount"
    output_ws['G5'] = data.total_amount
    output_ws['G5'].number_format =  u'"$ "#,##0.00'

    #Create top border and both borders setting
    top_border = Border(top=Side(style='thin'))
    bottom_border = Border(bottom=Side(style='thin'))
    both_border = Border(top=Side(style='thin'), bottom=Side(style='thin'))
    def apply_border_at_row (row, col_start, col_end, border_style):
        for c_idx in range(col_start,col_end + 1):
            output_ws.cell(row, c_idx).border=border_style

    #Offset and index
    output_row_offset = 9
    output_r_idx = output_row_offset

    #Heading of the table
    output_ws.cell(output_r_idx,1).value = "Date"
    output_ws.cell(output_r_idx,6).value = "Currency"
    output_ws.cell(output_r_idx,7).value = "Qty"
    output_ws.cell(output_r_idx,8).value = "Amount"
    #Apply right alignment on the number formatted column
    output_ws.cell(output_r_idx,6).alignment = Alignment(horizontal='right')
    output_ws.cell(output_r_idx,7).alignment = Alignment(horizontal='right')
    output_ws.cell(output_r_idx,8).alignment = Alignment(horizontal='right')
    #Apply the border setting to the heading row
    apply_border_at_row(output_r_idx, 1, 8, both_border)
    output_r_idx += 1

    #Go through all the data
    page_idx = 0
    max_page_line = 45
    for date in sorted(data.deposit_data.keys(), reverse = False):
        #Apply heading of the table for each page
        if (output_r_idx >= page_idx * max_page_line + 33):
            output_r_idx += max_page_line - 33
            page_idx += 1
            #Heading of the table
            output_ws.cell(output_r_idx,1).value = "Date"
            output_ws.cell(output_r_idx,6).value = "Currency"
            output_ws.cell(output_r_idx,7).value = "Qty"
            output_ws.cell(output_r_idx,8).value = "Amount"
            #Apply right alignment on the number formatted column
            output_ws.cell(output_r_idx,6).alignment = Alignment(horizontal='right')
            output_ws.cell(output_r_idx,7).alignment = Alignment(horizontal='right')
            output_ws.cell(output_r_idx,8).alignment = Alignment(horizontal='right')
            #Apply the border setting to the heading row
            apply_border_at_row(output_r_idx, 1, 8, both_border)
            output_r_idx += 1
        #Add each entry
        output_ws.cell(output_r_idx, 1).value = date.strftime("%a %d/%m/%y")
        output_ws.cell(output_r_idx, 1).alignment = Alignment(horizontal='left')
        apply_border_at_row(output_r_idx, 1, 8, top_border)
        #Write the currency, qty and amount for non-zero entries
        for cash_amount in data.cash_list:
            if (data.deposit_data[date][cash_amount]["n_amount"] > 0):
                #Write the cell in the excel file
                output_ws.cell(output_r_idx, 6).value = float(cash_amount)
                output_ws.cell(output_r_idx, 6).number_format = u'"$ "#,##0.00'
                output_ws.cell(output_r_idx, 7).value = data.deposit_data[date][cash_amount]["n_amount"]
                output_ws.cell(output_r_idx, 8).value = data.deposit_data[date][cash_amount]["each_amount"]
                output_ws.cell(output_r_idx, 8).number_format = u'"$ "#,##0.00'
                output_r_idx+=1
                #Add to the total_cash_amount
                if cash_amount in data.total_cash_amount:
                    data.total_cash_amount[cash_amount]['n_amount'] += data.deposit_data[date][cash_amount]["n_amount"]
                    data.total_cash_amount[cash_amount]['each_amount'] += data.deposit_data[date][cash_amount]["each_amount"]
                else:
                    data.total_cash_amount[cash_amount]={}
                    data.total_cash_amount[cash_amount]['n_amount']=data.deposit_data[date][cash_amount]["n_amount"]
                    data.total_cash_amount[cash_amount]['each_amount']=data.deposit_data[date][cash_amount]["each_amount"]
        #Write the daily total
        output_ws.cell(output_r_idx, 5).value = "Total Amount"
        output_ws.cell(output_r_idx, 8).value = data.deposit_data[date]["total_daily_amount"]
        output_ws.cell(output_r_idx, 8).number_format = u'"$ "#,##0.00'
        apply_border_at_row(output_r_idx, 1, 8, both_border)
        output_r_idx+=1
    #Apply heading of the table for each page
    if (output_r_idx >= page_idx * max_page_line + 33):
        output_r_idx += max_page_line - 33
        page_idx += 1
        #Heading of the table
        output_ws.cell(output_r_idx,1).value = "Date"
        output_ws.cell(output_r_idx,6).value = "Currency"
        output_ws.cell(output_r_idx,7).value = "Qty"
        output_ws.cell(output_r_idx,8).value = "Amount"
        #Apply right alignment on the number formatted column
        output_ws.cell(output_r_idx,6).alignment = Alignment(horizontal='right')
        output_ws.cell(output_r_idx,7).alignment = Alignment(horizontal='right')
        output_ws.cell(output_r_idx,8).alignment = Alignment(horizontal='right')
        #Apply the border setting to the heading row
        apply_border_at_row(output_r_idx, 1, 8, both_border)
        output_r_idx += 1
    #Write the total amount
    output_ws.cell(output_r_idx, 1).value = "Total"
    output_ws.cell(output_r_idx, 1).alignment = Alignment(horizontal='left')
    apply_border_at_row(output_r_idx, 1, 8, top_border)
    for cash_amount in data.cash_list:
        if cash_amount in data.total_cash_amount:
            #Write the cell in the excel file
            output_ws.cell(output_r_idx, 6).value = float(cash_amount)
            output_ws.cell(output_r_idx, 6).number_format = u'"$ "#,##0.00'
            output_ws.cell(output_r_idx, 7).value = data.total_cash_amount[cash_amount]["n_amount"]
            output_ws.cell(output_r_idx, 8).value = data.total_cash_amount[cash_amount]["each_amount"]
            output_ws.cell(output_r_idx, 8).number_format = u'"$ "#,##0.00'
            output_r_idx+=1
    output_ws.cell(output_r_idx, 5).value = "Total Amount"
    output_ws.cell(output_r_idx, 8).value = data.total_amount
    output_ws.cell(output_r_idx, 8).number_format = u'"$ "#,##0.00'
    apply_border_at_row(output_r_idx, 1, 8, both_border)


    ##Create a worksheet for quickbooks input
    #Create a new worksheet
    qb_title = f"QB {banking_date}"

    #Create worksheet for the daily cash amount
    output_ws = wb.create_sheet(qb_title)
    output_ws['A1'] = "Cash Deposit QB Breakdown"
    output_ws.column_dimensions['A'].width = 17
    output_ws.column_dimensions['B'].width = 11
    output_ws['A1'].font = Font(bold=True)

    #Add banking details
    output_ws['A3'] = "Banking date"
    output_ws['B3'] = data.banking_date
    output_ws['B3'].number_format = 'dd/mm/yyyy'
    output_ws['A4'] = "Start period"
    output_ws['B4'] = data.start_date
    output_ws['B4'].number_format = 'dd/mm/yyyy'
    output_ws['A5'] = "End period"
    output_ws['B5'] = data.end_date
    output_ws['B5'].number_format = 'dd/mm/yyyy'

    output_ws['E3'] = "Cash amount"
    output_ws['G3'] = data.cash_amount
    output_ws['G3'].number_format =  u'"$ "#,##0.00'
    output_ws['E4'] = "Coin amount"
    output_ws['G4'] = data.coin_amount
    output_ws['G4'].number_format =  u'"$ "#,##0.00'
    output_ws['E5'] = "Total amount"
    output_ws['G5'] = data.total_amount
    output_ws['G5'].number_format =  u'"$ "#,##0.00'

    #Offset and index
    output_row_offset = 9
    output_r_idx = output_row_offset

    #Heading of the table
    output_ws.cell(output_r_idx,1).value = "Date"
    output_ws.cell(output_r_idx,2).value = "Cash"
    output_ws.cell(output_r_idx,4).value = "Canteen"
    output_ws.cell(output_r_idx,5).value = "Playing"
    output_ws.cell(output_r_idx,6).value = "Hall Hire"
    output_ws.cell(output_r_idx,7).value = "Membership"
    output_ws.cell(output_r_idx,8).value = "Restring"
    output_ws.cell(output_r_idx,9).value = "Racket Hire"
    output_ws.cell(output_r_idx,10).value = "Shuttle"
    output_ws.cell(output_r_idx,11).value = "Tournament"
    output_ws.cell(output_r_idx,12).value = "Points"
    output_ws.cell(output_r_idx,13).value = "Locker"
    output_ws.cell(output_r_idx,14).value = "Misc"
    output_ws.cell(output_r_idx,15).value = "Total"
    output_r_idx+=1

    #Put in the total amount for each day
    for date in sorted(data.deposit_data.keys(), reverse = False):
        #Add each entry
        output_ws.cell(output_r_idx, 1).value = date.strftime("%a %d/%m/%y")
        output_ws.cell(output_r_idx, 1).alignment = Alignment(horizontal='left')
        #Get the detail in each amount
        output_ws.cell(output_r_idx, 2).value = data.deposit_data[date]["total_daily_amount"]
        output_ws.cell(output_r_idx, 2).number_format = u'"$ "#,##0.00'
        #Add the total
        output_ws.cell(output_r_idx, 15).value = f"=SUM(D{output_r_idx}:N{output_r_idx})"
        output_ws.cell(output_r_idx, 15).number_format = u'"$ "#,##0.00'
        output_r_idx+=1
    #Add the total
    output_ws.cell(output_r_idx, 1).value = "Total"
    output_ws.cell(output_r_idx, 2).value = f"=SUM(B{output_row_offset+1}:B{output_r_idx-1})"
    output_ws.cell(output_r_idx, 2).number_format = u'"$ "#,##0.00'

    #Save the spreadsheet
    generated_excel_filename = f"{os.getcwd()}\Report\CashDeposit 2023 - GUI - {banking_date}"
    wb.save(f"{generated_excel_filename}.xlsx")
    wb.close()

    # Write pdf
    generate_pdf(os.path.abspath(generated_excel_filename), f"Summary {banking_date}")

if __name__ == "__main__":
    # Assign data to struct
    report_banking_date = date.today().strftime("%A, %d/%m/%Y")
    report_start_date = date.today().strftime("%A, %d/%m/%Y")
    report_end_date = (date.today()+timedelta(days=1)).strftime("%A, %d/%m/%Y")
    report_date = (report_banking_date, report_start_date, report_end_date)

    report_cash_amount = 1500
    report_coin_amount = 1.5
    report_total_amount = 1501.5
    report_amount = (report_cash_amount, report_coin_amount, report_total_amount)

    cash_list = ["100", "50", "20", "10", "5", "2", "1", "0.50", "0.20", "0.10", "0.05"]
    deposit_data={}

    report_data = cash_deposit_struct(report_date, report_amount, deposit_data, cash_list)

    generate_report(report_data)
    print(f"Report Generated")