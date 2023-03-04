from openpyxl import load_workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment, Font
from datetime import datetime

def generate_report(filename, sheet_name):
    wb = load_workbook(f"{filename}.xlsx", data_only=True)
    ws = wb[f"{sheet_name}"]

    #Read data from the workbook
    max_row = len(ws['A'])

    #Create a new worksheet
    banking_date = ws.cell(1, 2).value
    banking_date_str = str(banking_date)
    banking_date_str = datetime.strptime(banking_date_str,'%Y-%m-%d %H:%M:%S').strftime('%d%b%y')
    summary_title = f"Summary {banking_date_str}"

    #Create worksheet for the daily cash amount
    output_ws = wb.create_sheet(summary_title)
    output_ws['A1'] = "Cash Deposit Report Summary"
    output_ws.column_dimensions['A'].width = 17
    output_ws.column_dimensions['B'].width = 11
    output_ws['A1'].font = Font(bold=True)

    #Add banking details
    output_ws['A3'] = "Banking date"
    output_ws['B3'] = ws.cell(1, 2).value
    output_ws['B3'].number_format = 'dd/mm/yyyy'
    output_ws['A4'] = "Start period"
    output_ws['B4'] = ws.cell(2, 2).value
    output_ws['B4'].number_format = 'dd/mm/yyyy'
    output_ws['A5'] = "End period"
    output_ws['B5'] = ws.cell(3, 2).value
    output_ws['B5'].number_format = 'dd/mm/yyyy'

    output_ws['E3'] = "Cash amount"
    output_ws['G3'] = ws.cell(5, 2).value
    output_ws['G3'].number_format =  u'"$ "#,##0.00'
    output_ws['E4'] = "Coin amount"
    output_ws['G4'] = ws.cell(6, 2).value
    output_ws['G4'].number_format =  u'"$ "#,##0.00'
    output_ws['E5'] = "Total amount"
    output_ws['G5'] = ws.cell(7, 2).value
    output_ws['G5'].number_format =  u'"$ "#,##0.00'

    #Create top border and both borders setting
    top_border = Border(top=Side(style='thin'))
    bottom_border = Border(bottom=Side(style='thin'))
    both_border = Border(top=Side(style='thin'), bottom=Side(style='thin'))
    def apply_border_at_row (row, col_start, col_end, border_style):
        for c_idx in range(col_start,col_end + 1):
            output_ws.cell(row, c_idx).border=border_style

    #Offset and index
    row_offset = 10
    col_offset = 2
    col_length = 12
    output_row_offset = 9
    data_len = max_row - row_offset + 1
    output_r_idx = output_row_offset

    #Heading of the table
    output_ws.cell(output_r_idx,1).value = "Date"
    output_ws.cell(output_r_idx,6).value = "Currency"
    output_ws.cell(output_r_idx,7).value = "Qty"
    output_ws.cell(output_r_idx,8).value = "Amount"
    #Apply right alignment on the number formatted column
    output_ws.cell(output_r_idx,6).alignment = Alignment(horizontal='right')
    output_ws.cell(output_r_idx,7).alignment = Alignment(horizontal='right')
    output_ws.cell(output_r_idx,8).alignment = Alignment(horizontal='right')
    #Apply the border setting to the heading row
    apply_border_at_row(output_r_idx, 1, 8, both_border)
    output_r_idx += 1


    #The rest of the record
    #Go through all the data in the input worksheet
    page_idx = 0
    max_page_line = 45
    for r_idx in range(data_len):
        #Apply heading of the table for each page
        if (output_r_idx > page_idx * max_page_line + 33):
            output_r_idx += max_page_line - 33
            page_idx += 1
            #Heading of the table
            output_ws.cell(output_r_idx,1).value = "Date"
            output_ws.cell(output_r_idx,2).value = "Day"
            output_ws.cell(output_r_idx,7).value = "Cash Total"
            #Apply right alignment on the number formatted column
            output_ws.cell(output_r_idx,6).alignment = Alignment(horizontal='right')
            output_ws.cell(output_r_idx,7).alignment = Alignment(horizontal='right')
            output_ws.cell(output_r_idx,8).alignment = Alignment(horizontal='right')
            #Apply the border setting to the heading row
            apply_border_at_row(output_r_idx, 1, 8, both_border)
            output_r_idx += 1
        #Add each entry
        output_ws.cell(output_r_idx, 1).value = ws.cell(row_offset + r_idx, 1).value
        date_entry_str = str(ws.cell(row_offset + r_idx, 1).value)
        if date_entry_str != "Blank" and date_entry_str != "Total":
            output_ws.cell(output_r_idx, 1).value = datetime.strptime(date_entry_str,'%Y-%m-%d %H:%M:%S').strftime('%a %x')
        # output_ws.cell(output_r_idx, 1).number_format = 'dd/mm/yyyy'
        output_ws.cell(output_r_idx, 1).alignment = Alignment(horizontal='left')
        apply_border_at_row(output_r_idx, 1, 8, top_border)
        #Go through the columns
        for c_idx in range(col_length):
            #Check if there is any value in any of the column in the row
            if ws.cell(row_offset + r_idx, col_offset + c_idx).value != None:
                #Get the detail in each amount
                if c_idx < col_length-1:
                    output_ws.cell(output_r_idx, 6).value = ws.cell(9, col_offset + c_idx).value
                    output_ws.cell(output_r_idx, 6).number_format = u'"$ "#,##0.00'
                    output_ws.cell(output_r_idx, 7).value = ws.cell(row_offset + r_idx, col_offset + c_idx).value
                    output_ws.cell(output_r_idx, 8).value = output_ws.cell(output_r_idx, 6).value * output_ws.cell(output_r_idx, 7).value
                    output_ws.cell(output_r_idx, 8).number_format = u'"$ "#,##0.00'
                else:
                    output_ws.cell(output_r_idx, 5).value = ws.cell(9, col_offset + c_idx).value
                    output_ws.cell(output_r_idx, 8).value = ws.cell(row_offset + r_idx, col_offset + c_idx).value
                    output_ws.cell(output_r_idx, 8).number_format = u'"$ "#,##0.00'
                    apply_border_at_row(output_r_idx, 1, 8, both_border)
                output_r_idx+=1


    ##Create a worksheet for quickbooks input
    #Create a new worksheet
    qb_title = f"QB {banking_date_str}"

    #Create worksheet for the daily cash amount
    output_ws = wb.create_sheet(qb_title)
    output_ws['A1'] = "Cash Deposit QB Breakdown"
    output_ws.column_dimensions['A'].width = 17
    output_ws.column_dimensions['B'].width = 11
    output_ws['A1'].font = Font(bold=True)

    #Add banking details
    output_ws['A3'] = "Banking date"
    output_ws['B3'] = ws.cell(1, 2).value
    output_ws['B3'].number_format = 'dd/mm/yyyy'
    output_ws['A4'] = "Start period"
    output_ws['B4'] = ws.cell(2, 2).value
    output_ws['B4'].number_format = 'dd/mm/yyyy'
    output_ws['A5'] = "End period"
    output_ws['B5'] = ws.cell(3, 2).value
    output_ws['B5'].number_format = 'dd/mm/yyyy'

    output_ws['E3'] = "Cash amount"
    output_ws['G3'] = ws.cell(5, 2).value
    output_ws['G3'].number_format =  u'"$ "#,##0.00'
    output_ws['E4'] = "Coin amount"
    output_ws['G4'] = ws.cell(6, 2).value
    output_ws['G4'].number_format =  u'"$ "#,##0.00'
    output_ws['E5'] = "Total amount"
    output_ws['G5'] = ws.cell(7, 2).value
    output_ws['G5'].number_format =  u'"$ "#,##0.00'

    #Offset and index
    row_offset = 10
    col_offset = 2
    col_length = 12
    output_row_offset = 9
    data_len = max_row - row_offset + 1
    output_r_idx = output_row_offset

    #Heading of the table
    output_ws.cell(output_r_idx,1).value = "Date"
    output_ws.cell(output_r_idx,2).value = "Cash"
    output_ws.cell(output_r_idx,4).value = "Social"
    output_ws.cell(output_r_idx,5).value = "Junior"
    output_ws.cell(output_r_idx,6).value = "Locker"
    output_ws.cell(output_r_idx,7).value = "Membership"
    output_ws.cell(output_r_idx,8).value = "Canteen"
    output_ws.cell(output_r_idx,9).value = "Hall Hire" 
    output_ws.cell(output_r_idx,10).value = "Total"
    output_r_idx+=1

    #The rest of the record
    #Go through all the data in the input worksheet
    page_idx = 0
    max_page_line = 45
    total_cash = 0
    for r_idx in range(data_len):
        #Add each entry
        output_ws.cell(output_r_idx, 1).value = ws.cell(row_offset + r_idx, 1).value
        date_entry_str = str(ws.cell(row_offset + r_idx, 1).value)
        if date_entry_str != "Blank" and date_entry_str!= "Total":
            output_ws.cell(output_r_idx, 1).value = datetime.strptime(date_entry_str,'%Y-%m-%d %H:%M:%S').strftime('%a %x')
        # output_ws.cell(output_r_idx, 1).number_format = 'dd/mm/yyyy'
        output_ws.cell(output_r_idx, 1).alignment = Alignment(horizontal='left')
        #Get the detail in each amount
        output_ws.cell(output_r_idx, 2).value = ws.cell(row_offset + r_idx, col_length+1).value
        output_ws.cell(output_r_idx, 2).number_format = u'"$ "#,##0.00'
        total_cash+=output_ws.cell(output_r_idx, 2).value
        output_r_idx+=1
    #Add the total
    output_ws.cell(output_r_idx, 1).value = "Total"
    output_ws.cell(output_r_idx, 2).value = total_cash
    output_ws.cell(output_r_idx, 2).number_format = u'"$ "#,##0.00'

    #Rename input sheet name
    ws.title = f"Input {banking_date_str}"

    #Save the spreadsheet
    wb.save(f"{filename} - {banking_date_str}.xlsx")
    wb.close()

    #Return workbook name
    print(f"Returning {filename} - {banking_date_str}")
    return f"{filename} - {banking_date_str}"

if __name__ == "__main__":
    output_filename = generate_report("D:\Data\Personal\Documents\Others\HunterBadminton\Data\Bank 2023 - NewFormat", "Sheet1")
    print(f"output filename is {output_filename}")